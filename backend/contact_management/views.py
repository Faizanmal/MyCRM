from rest_framework import viewsets, status, permissions, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from django.contrib.auth import get_user_model
from django.db.models import Q
from django.http import HttpResponse
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import io
from .models import Contact, ContactGroup, ContactImport
from .serializers import (
    ContactSerializer, ContactCreateSerializer, ContactGroupSerializer,
    ContactImportSerializer, ContactBulkUpdateSerializer, ContactExportSerializer
)

User = get_user_model()


class ContactViewSet(viewsets.ModelViewSet):
    """Contact management viewset"""
    queryset = Contact.objects.all()
    serializer_class = ContactSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['first_name', 'last_name', 'email', 'company_name', 'phone']
    ordering_fields = ['first_name', 'last_name', 'email', 'company_name', 'created_at']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        if self.action == 'create':
            return ContactCreateSerializer
        return ContactSerializer
    
    def get_queryset(self):
        queryset = Contact.objects.all()
        
        # Filter by assigned user if not admin
        if self.request.user.role != 'admin':
            queryset = queryset.filter(
                Q(assigned_to=self.request.user) | Q(created_by=self.request.user)
            )
        
        # Apply additional filters
        contact_type = self.request.query_params.get('contact_type')
        if contact_type:
            queryset = queryset.filter(contact_type=contact_type)
        
        assigned_to = self.request.query_params.get('assigned_to')
        if assigned_to:
            queryset = queryset.filter(assigned_to_id=assigned_to)
        
        tags = self.request.query_params.get('tags')
        if tags:
            tag_list = tags.split(',')
            queryset = queryset.filter(tags__overlap=tag_list)
        
        return queryset
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @action(detail=False, methods=['post'])
    def bulk_update(self, request):
        """Bulk update contacts"""
        serializer = ContactBulkUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        contact_ids = serializer.validated_data['contact_ids']
        updates = serializer.validated_data['updates']
        
        # Check permissions
        contacts = Contact.objects.filter(id__in=contact_ids)
        if self.request.user.role != 'admin':
            contacts = contacts.filter(
                Q(assigned_to=self.request.user) | Q(created_by=self.request.user)
            )
        
        updated_count = contacts.update(**updates)
        
        return Response({
            'message': f'{updated_count} contacts updated successfully',
            'updated_count': updated_count
        })
    
    @action(detail=False, methods=['post'])
    def bulk_delete(self, request):
        """Bulk delete contacts"""
        contact_ids = request.data.get('contact_ids', [])
        
        if not contact_ids:
            return Response(
                {'error': 'No contact IDs provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check permissions
        contacts = Contact.objects.filter(id__in=contact_ids)
        if self.request.user.role != 'admin':
            contacts = contacts.filter(
                Q(assigned_to=self.request.user) | Q(created_by=self.request.user)
            )
        
        deleted_count = contacts.count()
        contacts.delete()
        
        return Response({
            'message': f'{deleted_count} contacts deleted successfully',
            'deleted_count': deleted_count
        })
    
    @action(detail=False, methods=['post'])
    def export(self, request):
        """Export contacts to various formats"""
        serializer = ContactExportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        format_type = serializer.validated_data['format']
        fields = serializer.validated_data.get('fields', [])
        filters = serializer.validated_data.get('filters', {})
        group_id = serializer.validated_data.get('group_id')
        
        # Get contacts based on filters
        queryset = self.get_queryset()
        
        if group_id:
            try:
                group = ContactGroup.objects.get(id=group_id)
                queryset = queryset.filter(id__in=group.contacts.values_list('id', flat=True))
            except ContactGroup.DoesNotExist:
                return Response(
                    {'error': 'Contact group not found'},
                    status=status.HTTP_404_NOT_FOUND
                )
        
        # Apply additional filters
        for key, value in filters.items():
            if hasattr(Contact, key):
                queryset = queryset.filter(**{key: value})
        
        if format_type == 'csv':
            return self._export_csv(queryset, fields)
        elif format_type == 'excel':
            return self._export_excel(queryset, fields)
        elif format_type == 'pdf':
            return self._export_pdf(queryset, fields)
    
    def _export_csv(self, queryset, fields):
        """Export contacts to CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="contacts.csv"'
        
        writer = csv.writer(response)
        
        # Write headers
        if fields:
            writer.writerow(fields)
        else:
            writer.writerow([
                'First Name', 'Last Name', 'Email', 'Phone', 'Company',
                'Job Title', 'Contact Type', 'Status', 'Created At'
            ])
        
        # Write data
        for contact in queryset:
            if fields:
                row = [getattr(contact, field, '') for field in fields]
            else:
                row = [
                    contact.first_name, contact.last_name, contact.email,
                    contact.phone, contact.company_name, contact.job_title,
                    contact.contact_type, contact.status, contact.created_at.strftime('%Y-%m-%d')
                ]
            writer.writerow(row)
        
        return response
    
    def _export_excel(self, queryset, fields):
        """Export contacts to Excel"""
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="contacts.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contacts"
        
        # Write headers
        if fields:
            headers = fields
        else:
            headers = [
                'First Name', 'Last Name', 'Email', 'Phone', 'Company',
                'Job Title', 'Contact Type', 'Status', 'Created At'
            ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Write data
        for row, contact in enumerate(queryset, 2):
            if fields:
                values = [getattr(contact, field, '') for field in fields]
            else:
                values = [
                    contact.first_name, contact.last_name, contact.email,
                    contact.phone, contact.company_name, contact.job_title,
                    contact.contact_type, contact.status, contact.created_at.strftime('%Y-%m-%d')
                ]
            
            for col, value in enumerate(values, 1):
                ws.cell(row=row, column=col).value = value
        
        wb.save(response)
        return response
    
    def _export_pdf(self, queryset, fields):
        """Export contacts to PDF"""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="contacts.pdf"'
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title = Paragraph("Contacts Export", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Table data
        if fields:
            headers = fields
        else:
            headers = [
                'First Name', 'Last Name', 'Email', 'Phone', 'Company',
                'Job Title', 'Contact Type', 'Status'
            ]
        
        data = [headers]
        
        for contact in queryset:
            if fields:
                row = [str(getattr(contact, field, '')) for field in fields]
            else:
                row = [
                    contact.first_name, contact.last_name, contact.email,
                    contact.phone, contact.company_name, contact.job_title,
                    contact.contact_type, contact.status
                ]
            data.append(row)
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        doc.build(story)
        
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        
        return response


class ContactGroupViewSet(viewsets.ModelViewSet):
    """Contact group management viewset"""
    queryset = ContactGroup.objects.all()
    serializer_class = ContactGroupSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return ContactGroup.objects.filter(created_by=self.request.user)
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def add_contacts(self, request, pk=None):
        """Add contacts to group"""
        group = self.get_object()
        contact_ids = request.data.get('contact_ids', [])
        
        contacts = Contact.objects.filter(id__in=contact_ids)
        group.contacts.add(*contacts)
        
        return Response({
            'message': f'{len(contact_ids)} contacts added to group',
            'contact_count': group.contacts.count()
        })
    
    @action(detail=True, methods=['post'])
    def remove_contacts(self, request, pk=None):
        """Remove contacts from group"""
        group = self.get_object()
        contact_ids = request.data.get('contact_ids', [])
        
        contacts = Contact.objects.filter(id__in=contact_ids)
        group.contacts.remove(*contacts)
        
        return Response({
            'message': f'{len(contact_ids)} contacts removed from group',
            'contact_count': group.contacts.count()
        })


class ContactImportViewSet(viewsets.ReadOnlyModelViewSet):
    """Contact import tracking viewset"""
    queryset = ContactImport.objects.all()
    serializer_class = ContactImportSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return ContactImport.objects.filter(created_by=self.request.user)