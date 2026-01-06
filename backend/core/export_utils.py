"""
Data Export Utilities

Provides export functionality for:
- CSV export
- Excel export
- PDF generation
- JSON export
- Scheduled exports
"""

import csv
import io
import json
import logging
from collections.abc import Callable
from dataclasses import dataclass
from datetime import datetime
from typing import Any

from django.db.models import Model, QuerySet
from django.http import HttpResponse, StreamingHttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated

logger = logging.getLogger(__name__)


@dataclass
class ExportColumn:
    """Configuration for an export column."""
    field: str
    header: str
    formatter: Callable[[Any], str] | None = None
    width: int = 15


class DataExporter:
    """
    Base class for data export functionality.
    """

    def __init__(
        self,
        queryset: QuerySet,
        columns: list[ExportColumn],
        filename: str = 'export'
    ):
        self.queryset = queryset
        self.columns = columns
        self.filename = filename
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    def get_value(self, obj: Model, column: ExportColumn) -> str:
        """Extract and format a value from an object."""
        value = obj
        for part in column.field.split('__'):
            if value is None:
                break
            if hasattr(value, part):
                value = getattr(value, part)
            elif isinstance(value, dict):
                value = value.get(part)
            else:
                value = None

        if column.formatter and value is not None:
            return column.formatter(value)

        if value is None:
            return ''
        if isinstance(value, bool):
            return 'Yes' if value else 'No'
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')

        return str(value)

    def get_rows(self) -> list[list[str]]:
        """Generate rows of data."""
        rows = []
        for obj in self.queryset:
            row = [self.get_value(obj, col) for col in self.columns]
            rows.append(row)
        return rows

    def get_headers(self) -> list[str]:
        """Get column headers."""
        return [col.header for col in self.columns]


class CSVExporter(DataExporter):
    """Export data to CSV format."""

    def export(self) -> HttpResponse:
        """Generate CSV response."""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{self.filename}_{self.timestamp}.csv"'

        writer = csv.writer(response)
        writer.writerow(self.get_headers())

        for row in self.get_rows():
            writer.writerow(row)

        return response

    def export_streaming(self) -> StreamingHttpResponse:
        """Generate streaming CSV response for large datasets."""
        def generate():
            buffer = io.StringIO()
            writer = csv.writer(buffer)

            # Write headers
            writer.writerow(self.get_headers())
            buffer.seek(0)
            yield buffer.read()
            buffer.truncate(0)
            buffer.seek(0)

            # Write rows
            for obj in self.queryset.iterator():
                row = [self.get_value(obj, col) for col in self.columns]
                writer.writerow(row)
                buffer.seek(0)
                yield buffer.read()
                buffer.truncate(0)
                buffer.seek(0)

        response = StreamingHttpResponse(generate(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{self.filename}_{self.timestamp}.csv"'
        return response


class JSONExporter(DataExporter):
    """Export data to JSON format."""

    def export(self) -> HttpResponse:
        """Generate JSON response."""
        data = []

        for obj in self.queryset:
            item = {}
            for col in self.columns:
                item[col.field] = self.get_value(obj, col)
            data.append(item)

        response = HttpResponse(
            json.dumps(data, indent=2, ensure_ascii=False),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.filename}_{self.timestamp}.json"'
        return response


class ExcelExporter(DataExporter):
    """Export data to Excel format."""

    def export(self) -> HttpResponse:
        """Generate Excel response."""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            logger.error("openpyxl not installed, falling back to CSV")
            return CSVExporter(self.queryset, self.columns, self.filename).export()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.filename[:31]  # Excel sheet name limit

        # Header styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        # Write headers
        for col_idx, col in enumerate(self.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col.header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col.width

        # Write data
        for row_idx, row_data in enumerate(self.get_rows(), 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Create response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.filename}_{self.timestamp}.xlsx"'
        return response


class PDFExporter(DataExporter):
    """Export data to PDF format."""

    def export(self) -> HttpResponse:
        """Generate PDF response."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import landscape, letter
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle
        except ImportError:
            logger.error("reportlab not installed, falling back to CSV")
            return CSVExporter(self.queryset, self.columns, self.filename).export()

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title = Paragraph(f"<b>{self.filename}</b>", styles['Title'])
        elements.append(title)

        # Table data
        data = [self.get_headers()] + self.get_rows()

        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0F4F8')),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F4F8')]),
        ]))
        elements.append(table)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{self.filename}_{self.timestamp}.pdf"'
        return response


# Pre-configured exports for common models
def get_lead_columns() -> list[ExportColumn]:
    """Get export columns for Lead model."""
    return [
        ExportColumn('id', 'ID'),
        ExportColumn('first_name', 'First Name'),
        ExportColumn('last_name', 'Last Name'),
        ExportColumn('email', 'Email', width=25),
        ExportColumn('phone', 'Phone'),
        ExportColumn('company', 'Company'),
        ExportColumn('status', 'Status'),
        ExportColumn('source', 'Source'),
        ExportColumn('score', 'Score'),
        ExportColumn('assigned_to__username', 'Assigned To'),
        ExportColumn('created_at', 'Created At', width=20),
    ]


def get_contact_columns() -> list[ExportColumn]:
    """Get export columns for Contact model."""
    return [
        ExportColumn('id', 'ID'),
        ExportColumn('first_name', 'First Name'),
        ExportColumn('last_name', 'Last Name'),
        ExportColumn('email', 'Email', width=25),
        ExportColumn('phone', 'Phone'),
        ExportColumn('company', 'Company'),
        ExportColumn('job_title', 'Job Title'),
        ExportColumn('created_at', 'Created At', width=20),
    ]


def get_opportunity_columns() -> list[ExportColumn]:
    """Get export columns for Opportunity model."""
    return [
        ExportColumn('id', 'ID'),
        ExportColumn('name', 'Name', width=25),
        ExportColumn('value', 'Value', formatter=lambda v: f"${v:,.2f}"),
        ExportColumn('stage', 'Stage'),
        ExportColumn('probability', 'Probability', formatter=lambda v: f"{v}%"),
        ExportColumn('expected_close_date', 'Expected Close'),
        ExportColumn('owner__username', 'Owner'),
        ExportColumn('created_at', 'Created At', width=20),
    ]


def get_task_columns() -> list[ExportColumn]:
    """Get export columns for Task model."""
    return [
        ExportColumn('id', 'ID'),
        ExportColumn('title', 'Title', width=30),
        ExportColumn('description', 'Description', width=40),
        ExportColumn('status', 'Status'),
        ExportColumn('priority', 'Priority'),
        ExportColumn('due_date', 'Due Date'),
        ExportColumn('assigned_to__username', 'Assigned To'),
        ExportColumn('created_at', 'Created At', width=20),
    ]


# Export API Views
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_leads(request):
    """Export leads to specified format."""
    from lead_management.models import Lead

    format_type = request.query_params.get('format', 'csv')
    queryset = Lead.objects.filter(created_by=request.user).select_related('assigned_to')

    # Apply filters
    status = request.query_params.get('status')
    if status:
        queryset = queryset.filter(status=status)

    columns = get_lead_columns()

    if format_type == 'json':
        return JSONExporter(queryset, columns, 'leads').export()
    elif format_type == 'xlsx':
        return ExcelExporter(queryset, columns, 'leads').export()
    elif format_type == 'pdf':
        return PDFExporter(queryset, columns, 'leads').export()
    else:
        return CSVExporter(queryset, columns, 'leads').export()


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_contacts(request):
    """Export contacts to specified format."""
    from contact_management.models import Contact

    format_type = request.query_params.get('format', 'csv')
    queryset = Contact.objects.filter(created_by=request.user)

    columns = get_contact_columns()

    if format_type == 'json':
        return JSONExporter(queryset, columns, 'contacts').export()
    elif format_type == 'xlsx':
        return ExcelExporter(queryset, columns, 'contacts').export()
    elif format_type == 'pdf':
        return PDFExporter(queryset, columns, 'contacts').export()
    else:
        return CSVExporter(queryset, columns, 'contacts').export()


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_opportunities(request):
    """Export opportunities to specified format."""
    from opportunity_management.models import Opportunity

    format_type = request.query_params.get('format', 'csv')
    queryset = Opportunity.objects.filter(owner=request.user).select_related('owner')

    columns = get_opportunity_columns()

    if format_type == 'json':
        return JSONExporter(queryset, columns, 'opportunities').export()
    elif format_type == 'xlsx':
        return ExcelExporter(queryset, columns, 'opportunities').export()
    elif format_type == 'pdf':
        return PDFExporter(queryset, columns, 'opportunities').export()
    else:
        return CSVExporter(queryset, columns, 'opportunities').export()


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_tasks(request):
    """Export tasks to specified format."""
    from task_management.models import Task

    format_type = request.query_params.get('format', 'csv')
    queryset = Task.objects.filter(created_by=request.user).select_related('assigned_to')

    columns = get_task_columns()

    if format_type == 'json':
        return JSONExporter(queryset, columns, 'tasks').export()
    elif format_type == 'xlsx':
        return ExcelExporter(queryset, columns, 'tasks').export()
    elif format_type == 'pdf':
        return PDFExporter(queryset, columns, 'tasks').export()
    else:
        return CSVExporter(queryset, columns, 'tasks').export()


# URL patterns to add to urls.py:
# path('export/leads/', export_leads, name='export-leads'),
# path('export/contacts/', export_contacts, name='export-contacts'),
# path('export/opportunities/', export_opportunities, name='export-opportunities'),
# path('export/tasks/', export_tasks, name='export-tasks'),
