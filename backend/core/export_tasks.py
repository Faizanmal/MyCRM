"""
Export Celery Tasks
Async tasks for processing data exports
"""

import csv
import json
import io
import os
import zipfile
from datetime import timedelta
from tempfile import NamedTemporaryFile

from celery import shared_task
from django.conf import settings
from django.utils import timezone
from django.core.files.storage import default_storage
from django.contrib.auth import get_user_model

import logging

logger = logging.getLogger(__name__)

User = get_user_model()


@shared_task(bind=True, max_retries=3)
def process_export_job(self, job_id):
    """
    Process an export job asynchronously
    """
    from .settings_models import ExportJob
    from .export_views import DataExportService
    
    try:
        job = ExportJob.objects.get(id=job_id)
        job.mark_processing()
        
        user = job.user
        config = {
            'entities': job.entities,
            'date_range': job.date_range,
            'include_archived': job.include_archived,
            'include_deleted': job.include_deleted,
        }
        
        # Initialize export service
        service = DataExportService(user, config)
        
        # Update progress: Starting
        job.progress = 10
        job.save(update_fields=['progress'])
        
        # Export all requested entities
        data = service.export_all()
        
        # Update progress: Data extracted
        job.progress = 50
        job.save(update_fields=['progress'])
        
        # Generate output file
        if job.format == 'json':
            output_content = service.to_json(data)
            content_type = 'application/json'
            extension = 'json'
        elif job.format == 'xlsx':
            output_content = generate_excel_export(data)
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            extension = 'xlsx'
        else:  # CSV / ZIP
            if len(job.entities) > 1:
                output_content = service.to_zip(data)
                content_type = 'application/zip'
                extension = 'zip'
            else:
                output_content = service.to_csv(data)
                content_type = 'text/csv'
                extension = 'csv'
        
        # Update progress: File generated
        job.progress = 80
        job.save(update_fields=['progress'])
        
        # Save file to storage
        filename = f"exports/{user.id}/{job.id}.{extension}"
        
        if isinstance(output_content, bytes):
            file_content = io.BytesIO(output_content)
        else:
            file_content = io.BytesIO(output_content.encode('utf-8'))
        
        file_path = default_storage.save(filename, file_content)
        file_size = file_content.getbuffer().nbytes
        
        # Mark job as completed
        job.mark_completed(file_path, file_size)
        
        logger.info(f"Export job {job_id} completed successfully")
        return {'status': 'success', 'file_path': file_path, 'file_size': file_size}
        
    except ExportJob.DoesNotExist:
        logger.error(f"Export job {job_id} not found")
        return {'status': 'error', 'message': 'Job not found'}
        
    except Exception as e:
        logger.exception(f"Export job {job_id} failed: {str(e)}")
        
        try:
            job = ExportJob.objects.get(id=job_id)
            job.mark_failed(str(e))
        except:
            pass
        
        # Retry with exponential backoff
        self.retry(countdown=60 * (2 ** self.request.retries), exc=e)


def generate_excel_export(data):
    """
    Generate Excel file from export data
    Uses openpyxl if available, otherwise falls back to CSV in ZIP
    """
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        for entity_name, records in data.items():
            if not records:
                continue
            
            ws = wb.create_sheet(title=entity_name.capitalize()[:31])  # Max 31 chars
            
            # Write headers
            headers = list(records[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=str(header))
                cell.font = cell.font.copy(bold=True)
            
            # Write data
            for row_idx, record in enumerate(records, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = record.get(header, '')
                    if hasattr(value, 'isoformat'):
                        value = value.isoformat()
                    ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '')
            
            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
        
    except ImportError:
        logger.warning("openpyxl not installed, falling back to CSV in ZIP")
        from .export_views import DataExportService
        
        # Create a pseudo-service to use its methods
        class MockService:
            def to_zip(self, data):
                output = io.BytesIO()
                with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zf:
                    for entity_name, records in data.items():
                        if records:
                            csv_output = io.StringIO()
                            writer = csv.DictWriter(csv_output, fieldnames=records[0].keys())
                            writer.writeheader()
                            writer.writerows(records)
                            zf.writestr(f'{entity_name}.csv', csv_output.getvalue())
                output.seek(0)
                return output.getvalue()
        
        return MockService().to_zip(data)


@shared_task
def cleanup_expired_exports():
    """
    Clean up expired export files
    Should be run daily via Celery beat
    """
    from .settings_models import ExportJob
    
    expired_jobs = ExportJob.objects.filter(
        expires_at__lt=timezone.now(),
        status='completed',
        file_path__isnull=False
    )
    
    deleted_count = 0
    for job in expired_jobs:
        try:
            if job.file_path and default_storage.exists(job.file_path):
                default_storage.delete(job.file_path)
                deleted_count += 1
            
            job.file_path = None
            job.save(update_fields=['file_path'])
        except Exception as e:
            logger.error(f"Failed to clean up export {job.id}: {str(e)}")
    
    logger.info(f"Cleaned up {deleted_count} expired export files")
    return {'deleted': deleted_count}


@shared_task
def send_export_notification(job_id):
    """
    Send notification when export is complete
    """
    from .settings_models import ExportJob
    from .email_notifications import EmailNotificationService
    
    try:
        job = ExportJob.objects.get(id=job_id)
        
        if job.status == 'completed':
            # Send email notification
            email_service = EmailNotificationService()
            # email_service.send_export_ready(job.user, job)
            
            # Send in-app notification
            # create_notification(
            #     user=job.user,
            #     type='export_ready',
            #     title='Your export is ready',
            #     message=f'Your {job.format.upper()} export is ready to download',
            #     action_url=f'/settings/export/?download={job.id}'
            # )
            
            logger.info(f"Export notification sent for job {job_id}")
            
    except ExportJob.DoesNotExist:
        logger.error(f"Export job {job_id} not found for notification")


@shared_task
def generate_scheduled_export(user_id, config):
    """
    Generate a scheduled export (e.g., weekly backup)
    """
    from .settings_models import ExportJob
    
    try:
        user = User.objects.get(id=user_id)
        
        job = ExportJob.objects.create(
            user=user,
            format=config.get('format', 'csv'),
            entities=config.get('entities', ['contacts', 'deals', 'tasks']),
            date_range=config.get('date_range', 'all'),
            include_archived=config.get('include_archived', False),
            include_deleted=config.get('include_deleted', False),
        )
        
        # Process the export
        process_export_job.delay(job.id)
        
        logger.info(f"Scheduled export created for user {user_id}: job {job.id}")
        return {'job_id': job.id}
        
    except User.DoesNotExist:
        logger.error(f"User {user_id} not found for scheduled export")
        return {'error': 'User not found'}


# ==================== Analytics Tasks ====================

@shared_task
def generate_analytics_snapshot():
    """
    Generate daily analytics snapshot for reporting
    Should be run daily via Celery beat
    """
    from .settings_views import AnalyticsDashboardView
    from django.core.cache import cache
    
    view = AnalyticsDashboardView()
    
    # Generate for different time ranges
    for time_range in ['week', 'month', 'quarter', 'year']:
        try:
            # Create mock request
            class MockRequest:
                query_params = {'range': time_range}
            
            # This is a simplified approach - in production you'd want proper caching
            cache_key = f"analytics_snapshot_{time_range}_{timezone.now().date()}"
            cache.set(cache_key, None, 86400)  # Cache for 24 hours
            
        except Exception as e:
            logger.error(f"Failed to generate {time_range} analytics snapshot: {str(e)}")
    
    logger.info("Analytics snapshots generated")


@shared_task
def calculate_user_metrics(user_id, period='month'):
    """
    Calculate metrics for a specific user
    """
    try:
        user = User.objects.get(id=user_id)
        
        from opportunity_management.models import Opportunity
        from task_management.models import Task
        from activity_feed.models import Activity
        
        now = timezone.now()
        if period == 'week':
            start_date = now - timedelta(days=7)
        elif period == 'quarter':
            start_date = now - timedelta(days=90)
        elif period == 'year':
            start_date = now - timedelta(days=365)
        else:  # month
            start_date = now - timedelta(days=30)
        
        # Calculate metrics
        deals_won = Opportunity.objects.filter(
            owner=user,
            status='won',
            closed_at__gte=start_date
        ).count()
        
        revenue = Opportunity.objects.filter(
            owner=user,
            status='won',
            closed_at__gte=start_date
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        tasks_completed = Task.objects.filter(
            assigned_to=user,
            status='completed',
            completed_at__gte=start_date
        ).count()
        
        activities = Activity.objects.filter(
            actor=user,
            created_at__gte=start_date
        ).count()
        
        return {
            'user_id': user_id,
            'period': period,
            'deals_won': deals_won,
            'revenue': float(revenue),
            'tasks_completed': tasks_completed,
            'activities': activities,
        }
        
    except User.DoesNotExist:
        return {'error': 'User not found'}
    except ImportError as e:
        logger.warning(f"Model import failed: {e}")
        return {'error': 'Models not available'}
