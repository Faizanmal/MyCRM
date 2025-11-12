"""
Data Import/Export Module
Handles bulk data operations with validation and mapping
"""

from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils import timezone
from django.apps import apps
import csv
import openpyxl
from io import BytesIO, StringIO
import pandas as pd
import logging
import json

User = get_user_model()
logger = logging.getLogger(__name__)


class DataImporter:
    """Handle data import from various formats"""
    
    SUPPORTED_FORMATS = ['csv', 'xlsx', 'xls', 'json']
    
    def __init__(self, model_name, file_obj, file_format, field_mapping=None, user=None):
        """
        Initialize data importer
        
        Args:
            model_name: Model to import data into (e.g., 'contact_management.Contact')
            file_obj: File object or file path
            file_format: File format ('csv', 'xlsx', 'json')
            field_mapping: Dict mapping file columns to model fields
            user: User performing the import
        """
        self.model_name = model_name
        self.file_obj = file_obj
        self.file_format = file_format.lower()
        self.field_mapping = field_mapping or {}
        self.user = user
        self.errors = []
        self.imported_count = 0
        self.skipped_count = 0
        
        # Get model class
        try:
            self.model = apps.get_model(model_name)
        except LookupError:
            raise ValueError(f"Model {model_name} not found")
    
    def import_data(self, validate_only=False, batch_size=100):
        """
        Import data from file
        
        Args:
            validate_only: If True, only validate without importing
            batch_size: Number of records to process per batch
        
        Returns:
            dict with import results
        """
        # Read data based on format
        if self.file_format == 'csv':
            data = self._read_csv()
        elif self.file_format in ['xlsx', 'xls']:
            data = self._read_excel()
        elif self.file_format == 'json':
            data = self._read_json()
        else:
            raise ValueError(f"Unsupported format: {self.file_format}")
        
        # Process data
        results = self._process_data(data, validate_only, batch_size)
        
        # Create import log
        from .models import DataImportLog
        DataImportLog.objects.create(
            model_name=self.model_name,
            file_format=self.file_format,
            total_records=len(data),
            imported_records=self.imported_count,
            skipped_records=self.skipped_count,
            errors=self.errors,
            field_mapping=self.field_mapping,
            validate_only=validate_only,
            imported_by=self.user,
            status='completed' if not self.errors else 'completed_with_errors'
        )
        
        return results
    
    def _read_csv(self):
        """Read CSV file"""
        if isinstance(self.file_obj, bytes):
            content = self.file_obj.decode('utf-8')
            reader = csv.DictReader(StringIO(content))
        else:
            reader = csv.DictReader(self.file_obj)
        
        return list(reader)
    
    def _read_excel(self):
        """Read Excel file"""
        if isinstance(self.file_obj, bytes):
            wb = openpyxl.load_workbook(BytesIO(self.file_obj))
        else:
            wb = openpyxl.load_workbook(self.file_obj)
        
        ws = wb.active
        data = []
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Read data rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))
        
        return data
    
    def _read_json(self):
        """Read JSON file"""
        if isinstance(self.file_obj, bytes):
            content = self.file_obj.decode('utf-8')
            return json.loads(content)
        else:
            return json.load(self.file_obj)
    
    def _process_data(self, data, validate_only, batch_size):
        """Process and import data"""
        records_to_create = []
        
        for idx, row in enumerate(data):
            try:
                # Map fields
                mapped_data = self._map_fields(row)
                
                # Validate data
                validation_errors = self._validate_record(mapped_data)
                if validation_errors:
                    self.errors.append({
                        'row': idx + 2,  # +2 for header and 0-index
                        'errors': validation_errors
                    })
                    self.skipped_count += 1
                    continue
                
                if not validate_only:
                    records_to_create.append(self.model(**mapped_data))
                
                # Batch create
                if len(records_to_create) >= batch_size:
                    self._bulk_create(records_to_create)
                    records_to_create = []
            
            except Exception as e:
                self.errors.append({
                    'row': idx + 2,
                    'errors': [str(e)]
                })
                self.skipped_count += 1
        
        # Create remaining records
        if records_to_create and not validate_only:
            self._bulk_create(records_to_create)
        
        return {
            'imported': self.imported_count,
            'skipped': self.skipped_count,
            'errors': self.errors,
            'total': len(data)
        }
    
    def _map_fields(self, row):
        """Map file fields to model fields"""
        mapped = {}
        
        for file_field, model_field in self.field_mapping.items():
            if file_field in row:
                value = row[file_field]
                # Clean and convert value
                mapped[model_field] = self._convert_value(model_field, value)
        
        # Add default values
        if self.user and hasattr(self.model, 'created_by'):
            mapped['created_by'] = self.user
        
        return mapped
    
    def _convert_value(self, field_name, value):
        """Convert value to appropriate type"""
        if value is None or value == '':
            return None
        
        # Get field type
        field = self.model._meta.get_field(field_name)
        
        # Convert based on field type
        if field.get_internal_type() == 'BooleanField':
            return str(value).lower() in ['true', '1', 'yes', 'y']
        elif field.get_internal_type() in ['IntegerField', 'BigIntegerField']:
            return int(value)
        elif field.get_internal_type() in ['FloatField', 'DecimalField']:
            return float(value)
        elif field.get_internal_type() in ['DateField', 'DateTimeField']:
            return pd.to_datetime(value)
        
        return str(value)
    
    def _validate_record(self, data):
        """Validate record data"""
        errors = []
        
        # Check required fields
        for field in self.model._meta.fields:
            if not field.null and not field.blank and field.name not in data:
                if not field.has_default():
                    errors.append(f"Required field '{field.name}' is missing")
        
        # Try to create instance to validate
        try:
            instance = self.model(**data)
            instance.full_clean()
        except Exception as e:
            errors.append(str(e))
        
        return errors
    
    @transaction.atomic
    def _bulk_create(self, records):
        """Bulk create records"""
        try:
            self.model.objects.bulk_create(records, ignore_conflicts=False)
            self.imported_count += len(records)
            logger.info(f"Imported {len(records)} {self.model_name} records")
        except Exception as e:
            logger.error(f"Error bulk creating records: {str(e)}")
            raise


class DataExporter:
    """Handle data export to various formats"""
    
    SUPPORTED_FORMATS = ['csv', 'xlsx', 'json']
    
    def __init__(self, model_name, queryset=None, fields=None, file_format='csv'):
        """
        Initialize data exporter
        
        Args:
            model_name: Model to export data from
            queryset: QuerySet to export (None = all records)
            fields: List of fields to export (None = all fields)
            file_format: Export format ('csv', 'xlsx', 'json')
        """
        self.model_name = model_name
        self.model = apps.get_model(model_name)
        self.queryset = queryset if queryset is not None else self.model.objects.all()
        self.fields = fields or self._get_exportable_fields()
        self.file_format = file_format.lower()
    
    def export(self):
        """
        Export data
        
        Returns:
            BytesIO object containing exported data
        """
        if self.file_format == 'csv':
            return self._export_csv()
        elif self.file_format == 'xlsx':
            return self._export_excel()
        elif self.file_format == 'json':
            return self._export_json()
        else:
            raise ValueError(f"Unsupported format: {self.file_format}")
    
    def _get_exportable_fields(self):
        """Get list of exportable fields"""
        fields = []
        for field in self.model._meta.fields:
            if not field.name.endswith('_id') and field.name not in ['password']:
                fields.append(field.name)
        return fields
    
    def _export_csv(self):
        """Export to CSV"""
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=self.fields)
        
        writer.writeheader()
        for obj in self.queryset:
            row = {}
            for field in self.fields:
                value = getattr(obj, field)
                row[field] = self._format_value(value)
            writer.writerow(row)
        
        # Convert to bytes
        output.seek(0)
        return BytesIO(output.getvalue().encode('utf-8'))
    
    def _export_excel(self):
        """Export to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.model._meta.verbose_name_plural[:31]  # Excel sheet name limit
        
        # Write header
        for col, field in enumerate(self.fields, start=1):
            ws.cell(row=1, column=col, value=field)
        
        # Write data
        for row_num, obj in enumerate(self.queryset, start=2):
            for col, field in enumerate(self.fields, start=1):
                value = getattr(obj, field)
                ws.cell(row=row_num, column=col, value=self._format_value(value))
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def _export_json(self):
        """Export to JSON"""
        data = []
        for obj in self.queryset:
            row = {}
            for field in self.fields:
                value = getattr(obj, field)
                row[field] = self._format_value(value)
            data.append(row)
        
        output = BytesIO()
        output.write(json.dumps(data, indent=2, default=str).encode('utf-8'))
        output.seek(0)
        return output
    
    def _format_value(self, value):
        """Format value for export"""
        if value is None:
            return ''
        if isinstance(value, (timezone.datetime, timezone.date)):
            return value.isoformat()
        return str(value)


class BulkOperationManager:
    """Handle bulk operations on records"""
    
    @staticmethod
    @transaction.atomic
    def bulk_update(model_name, record_ids, updates, user=None):
        """
        Bulk update records
        
        Args:
            model_name: Model name
            record_ids: List of record IDs to update
            updates: Dict of field: value to update
            user: User performing the operation
        
        Returns:
            Number of records updated
        """
        model = apps.get_model(model_name)
        
        records = model.objects.filter(id__in=record_ids)
        count = records.update(**updates)
        
        # Log audit trail
        from core.security import SecurityAuditLog
        if user:
            SecurityAuditLog.log_event(
                user=user,
                action='bulk_update',
                resource=f"{model_name}: {count} records",
                metadata={'updates': updates, 'record_ids': record_ids}
            )
        
        logger.info(f"Bulk updated {count} {model_name} records")
        return count
    
    @staticmethod
    @transaction.atomic
    def bulk_delete(model_name, record_ids, user=None):
        """
        Bulk delete records
        
        Args:
            model_name: Model name
            record_ids: List of record IDs to delete
            user: User performing the operation
        
        Returns:
            Number of records deleted
        """
        model = apps.get_model(model_name)
        
        count, _ = model.objects.filter(id__in=record_ids).delete()
        
        # Log audit trail
        from core.security import SecurityAuditLog
        if user:
            SecurityAuditLog.log_event(
                user=user,
                action='bulk_delete',
                resource=f"{model_name}: {count} records",
                metadata={'record_ids': record_ids},
                risk_level='high'
            )
        
        logger.info(f"Bulk deleted {count} {model_name} records")
        return count
    
    @staticmethod
    @transaction.atomic
    def bulk_assign(model_name, record_ids, assigned_to_id, user=None):
        """
        Bulk assign records to a user
        
        Args:
            model_name: Model name
            record_ids: List of record IDs to assign
            assigned_to_id: User ID to assign to
            user: User performing the operation
        
        Returns:
            Number of records assigned
        """
        model = apps.get_model(model_name)
        
        if not hasattr(model, 'assigned_to'):
            raise ValueError(f"{model_name} does not support assignment")
        
        count = model.objects.filter(id__in=record_ids).update(
            assigned_to_id=assigned_to_id
        )
        
        # Log audit trail
        from core.security import SecurityAuditLog
        if user:
            SecurityAuditLog.log_event(
                user=user,
                action='bulk_assign',
                resource=f"{model_name}: {count} records",
                metadata={'assigned_to_id': assigned_to_id, 'record_ids': record_ids}
            )
        
        logger.info(f"Bulk assigned {count} {model_name} records to user {assigned_to_id}")
        return count
